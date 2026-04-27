from __future__ import annotations

import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(
    "/Users/bokrantz/Documents/01. Chalmers/01. Projekt/01. Aktiva projekt/"
    "MATTER@SCALE/Research studies/WP1 Human/M@S WP1 - KSAOs and Capabilities.xlsx"
)
OUTPUT = Path(__file__).resolve().parents[1] / "data.js"

KSAO_TYPES = {"Knowledge", "Skills", "Abilities", "Other characteristics"}
TYPE_LABEL = {
    "Knowledge": "Knowledge",
    "Skills": "Skill",
    "Abilities": "Ability",
    "Other characteristics": "Other characteristics",
}


def clean(value):
    if value is None:
        return ""
    return str(value).strip()


def read_ksaos(workbook):
    sheet = workbook["KSAOs"]
    items = {}
    current_type = None

    for row in sheet.iter_rows(values_only=True):
        first = clean(row[0])
        if not first:
            continue
        if first in KSAO_TYPES:
            current_type = TYPE_LABEL[first]
            continue
        if current_type:
            items[first] = {
                "name": first,
                "type": current_type,
                "definition": clean(row[1]),
                "exemplar": clean(row[2]).replace("“", '"').replace("”", '"'),
                "roles": [],
                "sectors": [],
                "capabilities": [],
            }

    return items


def apply_matrix(workbook, sheet_name, target_key, items):
    sheet = workbook[sheet_name]
    headers = []
    current_type = None

    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        first = clean(row[0])
        if not first:
            continue
        if first in KSAO_TYPES:
            current_type = TYPE_LABEL[first]
            headers = [clean(cell) for cell in row[1:] if clean(cell)]
            continue
        if first not in items:
            items[first] = {
                "name": first,
                "type": current_type or "",
                "definition": "",
                "exemplar": "",
                "roles": [],
                "sectors": [],
                "capabilities": [],
            }
        selected = []
        for header, cell in zip(headers, row[1:]):
            if clean(cell).upper() == "X":
                selected.append(header)
        items[first][target_key] = selected


def read_capabilities(workbook, items):
    sheet = workbook["Capabilites"]
    capability_type = None
    capability_name = None
    capabilities = []
    by_name = {}
    capability_types = {"Operational capabilites": "Operational", "Dynamic capabilites": "Dynamic"}

    for row in sheet.iter_rows(values_only=True):
        first = clean(row[0])
        second = clean(row[1])
        if not first:
            continue
        if first in capability_types:
            capability_type = capability_types[first]
            capability_name = None
            continue
        if first == "KSAO":
            continue
        if second == "":
            capability_name = first
            by_name[capability_name] = {
                "name": capability_name,
                "type": capability_type,
                "ksaos": [],
            }
            capabilities.append(by_name[capability_name])
            continue
        if capability_name:
            by_name[capability_name]["ksaos"].append(first)
            if first in items:
                items[first]["capabilities"].append(
                    {"name": capability_name, "type": capability_type}
                )

    return capabilities


def enrich(items, capabilities):
    roles = ["Technician", "Supervisor", "Engineer", "Manager"]
    sectors = ["Automotive", "Battery"]
    types = ["Knowledge", "Skill", "Ability", "Other characteristics"]

    by_capability = {cap["name"]: cap for cap in capabilities}
    for item in items.values():
        item["capabilityNames"] = [cap["name"] for cap in item["capabilities"]]
        item["capabilityTypes"] = sorted({cap["type"] for cap in item["capabilities"]})
        item["searchText"] = " ".join(
            [
                item["name"],
                item["type"],
                item["definition"],
                item["exemplar"],
                " ".join(item["roles"]),
                " ".join(item["sectors"]),
                " ".join(item["capabilityNames"]),
            ]
        ).lower()

    role_counts = defaultdict(int)
    sector_counts = defaultdict(int)
    type_counts = defaultdict(int)
    for item in items.values():
        type_counts[item["type"]] += 1
        for role in item["roles"]:
            role_counts[role] += 1
        for sector in item["sectors"]:
            sector_counts[sector] += 1

    for cap in capabilities:
        cap_items = [items[name] for name in cap["ksaos"] if name in items]
        cap["roleCounts"] = {
            role: sum(1 for item in cap_items if role in item["roles"]) for role in roles
        }
        cap["sectorCounts"] = {
            sector: sum(1 for item in cap_items if sector in item["sectors"])
            for sector in sectors
        }
        cap["typeCounts"] = {
            item_type: sum(1 for item in cap_items if item["type"] == item_type)
            for item_type in types
        }
        cap["description"] = describe_capability(cap["name"], cap["type"])

    return {
        "source": WORKBOOK.name,
        "roles": roles,
        "sectors": sectors,
        "types": types,
        "ksaos": sorted(items.values(), key=lambda item: (types.index(item["type"]), item["name"])),
        "capabilities": capabilities,
        "summary": {
            "ksaoCount": len(items),
            "capabilityCount": len(capabilities),
            "roleCounts": dict(role_counts),
            "sectorCounts": dict(sector_counts),
            "typeCounts": dict(type_counts),
        },
    }


def describe_capability(name, capability_type):
    operational = {
        "Execution": "Day-to-day ability to restore, inspect, maintain, and commission technical systems.",
        "Planning": "Ability to translate maintenance goals into plans, resources, schedules, budgets, and programs.",
        "Controlling": "Ability to govern safety, quality, documentation, performance, risk, and information integrity.",
        "Improving": "Ability to remove recurring losses, improve methods, and raise technical and organizational performance.",
    }
    dynamic = {
        "Building": "Ability to develop people, teams, and future competence inside the maintenance organization.",
        "Integrating": "Ability to coordinate knowledge and action across internal functions, suppliers, and networks.",
        "Reconfiguring": "Ability to reshape assets, concepts, and ways of working when technology or production needs change.",
    }
    return (operational if capability_type == "Operational" else dynamic).get(name, "")


def main():
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    items = read_ksaos(workbook)
    apply_matrix(workbook, "KSAOs per role", "roles", items)
    apply_matrix(workbook, "KSAOs per sector", "sectors", items)
    capabilities = read_capabilities(workbook, items)
    payload = enrich(items, capabilities)
    OUTPUT.write_text(
        "window.KSAO_DATA = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    print(f"Wrote {OUTPUT} with {len(payload['ksaos'])} KSAOs")


if __name__ == "__main__":
    main()
